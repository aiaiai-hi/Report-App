import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
import io
from pathlib import Path
import warnings
from openpyxl import Workbook
from openpyxl.styles import Font
warnings.filterwarnings('ignore')

# Конфигурация страницы
st.set_page_config(
    page_title="Система управления отчетами",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Инициализация состояния сессии
if 'admin_mode' not in st.session_state:
    st.session_state.admin_mode = False

# CSS стили для улучшения внешнего вида
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .page-header {
        font-size: 2rem;
        font-weight: bold;
        color: #333;
        border-bottom: 2px solid #1f77b4;
        padding-bottom: 0.5rem;
        margin-bottom: 1.5rem;
    }
    .sidebar-header {
        font-size: 1.2rem;
        font-weight: bold;
        color: #1f77b4;
        margin-bottom: 1rem;
    }
    .admin-badge {
        background-color: #ff4b4b;
        color: white;
        padding: 0.2rem 0.5rem;
        border-radius: 0.3rem;
        font-size: 0.8rem;
        font-weight: bold;
    }
    .coming-soon {
        text-align: center;
        color: #666;
        font-style: italic;
        font-size: 1.2rem;
        margin-top: 3rem;
        padding: 2rem;
        border: 2px dashed #ccc;
        border-radius: 0.5rem;
    }

    # Добавьте эти стили к существующим CSS стилям
    
    .section-container {
        background-color: #f8f9fa;
        padding: 1.5rem;
        border-radius: 0.5rem;
        border-left: 4px solid #1f77b4;
        margin-bottom: 1rem;
    }
    
    .edit-mode-badge {
        background-color: #ff8c00;
        color: white;
        padding: 0.2rem 0.5rem;
        border-radius: 0.3rem;
        font-size: 0.7rem;
        font-weight: bold;
        margin-bottom: 0.5rem;
        display: inline-block;
    }
    
    .section-text {
        margin: 0;
        line-height: 1.6;
        font-size: 1rem;
    }
</style>
""", unsafe_allow_html=True)

# Класс ExcelTransformer из transform_test.py
class ExcelTransformer:
    def __init__(self, report_number=None):
        """
        Инициализация трансформера
        
        Args:
            report_number (str): Номер отчета для генерации кодов атрибутов
        """
        self.report_number = report_number or "R001"
        self.supported_extensions = ['.xlsx', '.xls', '.csv']
        self.report_types = ["Ручной", "Полуавтоматический", "Автоматический", "ИЛА"]
    
    def detect_data_type(self, values):
        """
        Автоматическое определение типа данных столбца
        
        Args:
            values: pandas Series с данными столбца
            
        Returns:
            str: тип данных ('текст', 'число', 'дата', 'флаг')
        """
        # Убираем пустые значения и NaN
        clean_values = values.dropna()
        if len(clean_values) == 0:
            return "текст"
        
        # Конвертируем в строки для анализа
        str_values = clean_values.astype(str).str.strip().str.lower()
        
        # Проверка на булевы значения (флаги)
        bool_indicators = {
            'да', 'нет', 'true', 'false', '1', '0', 'yes', 'no', 
            'y', 'n', 'вкл', 'выкл', 'on', 'off', 'активен', 'неактивен'
        }
        unique_values = set(str_values.unique())
        if unique_values.issubset(bool_indicators) and len(unique_values) <= 3:
            return "флаг"
        
        # Проверка на даты
        date_count = 0
        for val in clean_values:
            if self._is_date(val):
                date_count += 1
        
        if date_count / len(clean_values) > 0.7:  # 70% значений - даты
            return "дата"
        
        # Проверка на числа
        numeric_count = 0
        for val in clean_values:
            if self._is_numeric(val):
                numeric_count += 1
        
        if numeric_count / len(clean_values) > 0.8:  # 80% значений - числа
            return "число"
        
        return "текст"
    
    def _is_date(self, value):
        """Проверка, является ли значение датой"""
        if pd.isna(value):
            return False
            
        # Попробуем распарсить как дату
        date_formats = [
            '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y',
            '%d.%m.%y', '%d/%m/%y', '%y-%m-%d', '%d-%m-%y',
            '%Y.%m.%d', '%Y/%m/%d'
        ]
        
        str_val = str(value).strip()
        
        for fmt in date_formats:
            try:
                datetime.strptime(str_val, fmt)
                return True
            except ValueError:
                continue
                
        # Проверим pandas to_datetime
        try:
            pd.to_datetime(str_val, errors='raise')
            return True
        except:
            return False
    
    def _is_numeric(self, value):
        """Проверка, является ли значение числом"""
        if pd.isna(value):
            return False
            
        try:
            # Попробуем конвертировать в float
            float(str(value).replace(',', '.').replace(' ', ''))
            return True
        except ValueError:
            return False
    
    def load_from_uploaded_file(self, uploaded_file):
        """
        Загрузка данных из uploaded_file Streamlit
        
        Args:
            uploaded_file: файл, загруженный через st.file_uploader
            
        Returns:
            pandas.DataFrame: загруженные данные
        """
        try:
            file_extension = Path(uploaded_file.name).suffix.lower()
            
            if file_extension == '.csv':
                # Для CSV пробуем разные разделители и кодировки
                try:
                    df = pd.read_csv(uploaded_file, sep=',', encoding='utf-8')
                    if len(df.columns) > 1:
                        return df
                except:
                    pass
                
                try:
                    uploaded_file.seek(0)  # Сброс указателя файла
                    df = pd.read_csv(uploaded_file, sep=';', encoding='cp1251')
                    if len(df.columns) > 1:
                        return df
                except:
                    pass
                
                # Последняя попытка
                uploaded_file.seek(0)
                df = pd.read_csv(uploaded_file)
                
            else:
                # Для Excel файлов
                df = pd.read_excel(uploaded_file)
            
            return df
            
        except Exception as e:
            raise Exception(f"Ошибка при загрузке файла: {str(e)}")
    
    def transform_to_metadata(self, df, report_type):
        """
        Преобразование DataFrame в метаданные атрибутов
        
        Args:
            df (pandas.DataFrame): исходные данные
            report_type (str): тип отчета (Ручной, Полуавтоматический, Автоматический, ИЛА)
            
        Returns:
            pandas.DataFrame: метаданные атрибутов
        """
        metadata_list = []
        
        for idx, column in enumerate(df.columns, 1):
            # Получаем данные столбца
            column_data = df[column]
            
            # Определяем тип данных
            data_type = self.detect_data_type(column_data)
            
            # Определяем значения по умолчанию в зависимости от типа отчета
            if report_type in ["Ручной", "Полуавтоматический"]:
                tech_algorithm_to_be = "Ручной ввод"
                data_source_type = "Ручное заполнение"
            else:  # Автоматический или ИЛА
                tech_algorithm_to_be = ""
                data_source_type = "База данных"
            
            # Связь с ИС для ИЛА
            system_connection = "ИЛА One" if report_type == "ИЛА" else ""
            
            metadata_record = {
                'ReportCode_info': '',  # Будет заполнено позже
                'Noreportfield_info': idx,
                'name': column,
                'description': '',
                'TechAsIs': '',
                'BussAlgorythm': '',
                'TechAlgorythm': tech_algorithm_to_be,
                'algorithms_change_info': 'нет',
                'dbobjectlink': '',
                'base_type_info': data_source_type,
                'related_it_system_info': system_connection,
                'reportfields_codes': '',
                'reportfields_names': '',
                'reportfields_parent_term': '',
                'reportfields_domain': '',
                'required_attribute_info': 'да',
                'base_type_report_field': data_type,
                'base_calc_ref_ind_info': 'Базовый',
                'codeTable_info': '',
                'example': '',
                'isToDelete_info': ''
            }
            
            metadata_list.append(metadata_record)
        
        metadata_df = pd.DataFrame(metadata_list)
        
        # Заполняем код атрибута после создания DataFrame
        metadata_df['ReportCode_info'] = metadata_df['Noreportfield_info'].apply(
            lambda x: f"{self.report_number}_{x:03d}"
        )
        
        return metadata_df
    
    def create_excel_download(self, metadata_df):
        """
        Создание Excel файла для скачивания с заголовками и пользовательскими названиями
        
        Args:
            metadata_df (pandas.DataFrame): метаданные для сохранения
            
        Returns:
            bytes: данные Excel файла
        """
        output = io.BytesIO()
        
        # Создаем новую книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Атрибут отчета"
        
        # Технические заголовки (скрытая строка)
        technical_headers = [
            'ReportCode_info', 'Noreportfield_info', 'name', 'description', 'TechAsIs', 
            'BussAlgorythm', 'TechAlgorythm', 'algorithms_change_info', 'dbobjectlink', 
            'base_type_info', 'related_it_system_info', 'reportfields_codes', 
            'reportfields_names', 'reportfields_parent_term', 'reportfields_domain', 
            'required_attribute_info', 'base_type_report_field', 'base_calc_ref_ind_info', 
            'codeTable_info', 'example', 'isToDelete_info'
        ]
        
        # Пользовательские заголовки (видимая строка)
        user_headers = [
            'Код атрибута отчета', 
            '№ атрибута отчета', 
            'Наименование атрибута', 
            'Бизнес-алгоритм AS IS',
            'Технический алгоритм AS IS', 
            'Бизнес-алгоритм TO BE',
            'Технический алгоритм TO BE', 
            'Алгоритм изменен', 
            'Физические атрибуты', 
            'Тип источника данных',
            'Связь с информационной системой', 
            'Код термина/терминов',
            'Наименование термина/терминов', 
            'Наименование родительской сущности термина/терминов', 
            'Домен термина/терминов', 
            'Обязательный атрибут для заполнения', 
            'Базовый тип атрибута (Текст, Число, Дата, Флаг)', 
            'Признак атрибута (Базовый, Расчетный, Справочный)', 
            'Наименование справочника', 
            'Примечание', 
            'Помечен к удалению (да/нет)'
        ]
        
        # Записываем технические заголовки в первую строку (скрытую)
        for col_idx, header in enumerate(technical_headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Скрываем первую строку
        ws.row_dimensions[1].hidden = True
        
        # Записываем пользовательские заголовки во вторую строку (видимую)
        for col_idx, header in enumerate(user_headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            # Делаем заголовки полужирными
            cell.font = Font(bold=True)
        
        # Закрепляем первые две строки
        ws.freeze_panes = ws.cell(row=3, column=1)
        
        # Записываем данные начиная с третьей строки
        for row_idx, (_, row) in enumerate(metadata_df.iterrows(), 3):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Автоподбор ширины столбцов (учитываем все строки включая заголовки)
        for col_idx, column_letter in enumerate([chr(65 + i) for i in range(len(technical_headers))], 0):
            max_length = 0
            
            # Проверяем техническую строку
            if len(technical_headers) > col_idx:
                if len(str(technical_headers[col_idx])) > max_length:
                    max_length = len(str(technical_headers[col_idx]))
            
            # Проверяем пользовательскую строку
            if len(user_headers) > col_idx:
                if len(str(user_headers[col_idx])) > max_length:
                    max_length = len(str(user_headers[col_idx]))
            
            # Проверяем данные
            for row_idx in range(3, len(metadata_df) + 3):
                cell_value = ws.cell(row=row_idx, column=col_idx + 1).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в BytesIO
        wb.save(output)
        output.seek(0)
        return output.getvalue()

# Заголовок приложения
st.markdown('<div class="main-header">📊 Система управления отчетами</div>', unsafe_allow_html=True)

# Боковая панель с навигацией
with st.sidebar:
    st.markdown('<div class="sidebar-header">🧭 Навигация</div>', unsafe_allow_html=True)
    
    # Основное меню
    st.markdown("**Основные функции:**")
    main_pages = {
        "📋 Инструкции по отчетам": "instructions",
        "⚡ Действия с отчетами": "actions", 
        "🏷️ Сформировать атрибуты и термины": "attributes",
        "📈 Дашборд по актуальности отчетов": "dashboard",
        "🤖 Задать вопрос (ИИ ассистент)": "ai_assistant",
        "💬 Оставить обратную связь": "feedback"
    }
    
    # Для основного меню:
    if 'selected_page' not in st.session_state:
        st.session_state.selected_page = "📋 Инструкции по отчетам"
    
    for page_name in main_pages.keys():
        if st.button(page_name, key=f"btn_{main_pages[page_name]}", use_container_width=True):
            st.session_state.selected_page = page_name
    
    st.markdown("---")
    
    # Админ панель
    st.markdown("**Административная панель:**")
    admin_toggle = st.checkbox("🔐 Режим администратора", key="admin_toggle")
    st.session_state.admin_mode = admin_toggle
    
    if st.session_state.admin_mode:
        st.markdown('<span class="admin-badge">АДМИН РЕЖИМ</span>', unsafe_allow_html=True)
        st.markdown("")
        
        admin_pages = {
            "🔍 Контроль публикации отчетов": "admin_control",
            "📊 Статистика по публикации": "admin_stats", 
            "⚠️ Проблемные вопросы": "admin_issues"
        }
        
        # Для админ меню:
        if 'selected_admin_page' not in st.session_state:
            st.session_state.selected_admin_page = "🔍 Контроль публикации отчетов"
            
        for page_name in admin_pages.keys():
            if st.button(page_name, key=f"admin_btn_{admin_pages[page_name]}", use_container_width=True):
                st.session_state.selected_admin_page = page_name
    
    # Информация о системе
    st.markdown("---")
    st.markdown("**ℹ️ Информация о системе**")
    st.caption(f"Время: {datetime.now().strftime('%H:%M:%S')}")
    st.caption("Версия: 1.0.0")

# Функции для отображения страниц
def show_instructions():
    st.markdown('<div class="page-header">📋 Инструкции по отчетам</div>', unsafe_allow_html=True)
    st.markdown('<div class="coming-soon">🚧 Здесь будут размещены подробные инструкции по работе с отчетами</div>', unsafe_allow_html=True)

def show_actions():
    # st.markdown('<div class="page-header">⚡ Действия с отчетами</div>', unsafe_allow_html=True)
    # st.markdown('<div class="coming-soon">🚧 Здесь будут доступны различные действия с отчетами: создание, редактирование, удаление</div>', unsafe_allow_html=True)

    st.markdown('<div class="page-header">⚡ Действия с отчетами</div>', unsafe_allow_html=True)
    
    # Инициализация состояния для редактируемых текстов
    if 'action_texts' not in st.session_state:
        st.session_state.action_texts = {
            'register': 'Чтобы зарегистрировать новый отчет необходимо зайти в Бизнес-глоссарий и в левом меню найти раздел "Запросы - Отчеты". В правом верхнем углу выбрать кнопку "Создать".',
            'automate': 'Если у вас уже существует отчет, который собирается регулярно "вручную", проверьте наличие зарегистрированной информации об этом отчете в Бизнес-глоссарии. Если "ручной отчет" отсутствует в реестре отчетов, то прежде, чем направлять заявку на автоматизацию, необходимо пройти регистрацию "ручного отчета". После этого можете переходить к заявке на автоматизацию отчета. Автоматизация существующего ручного отчета/ автоматизация нового отчета ("ручной" отчет по аналогичной форме не собирается с ССП или РФ). Для подачи заявки на автоматизацию отчета, необходимо сформировать запрос в Бизнес-глоссарии: в левом меню найти раздел "Запросы - Отчеты". В правом верхнем углу выбрать кнопку "Создать".',
            'update': 'Чтобы актуализировать существующий отчет ...',
            'change_owner': 'Чтобы сменить владельца отчета, необходимо выбрать 1 из 2х вариантов: 1. Направить служебную записку в адрес ДБД в свободной форме или 2. Создать запрос на смену владельца отчета в Бизнес-глоссарии, предварительно уточнив ФИО ответственного, кто будет принимать отчет, чтобы указать его владельцем запроса после передачи отчета новому владельцу',
            'delete': 'Чтобы удалить отчет, необходимо сформировать запрос на удаление отчета. Если отчет автоматизирован, то необходимо приложить BIQ, уточнить ответственного за автоматизацию, чтобы передать данную информацию для отключения отчета в системе.'
        }
    
    # Разделы с действиями
    sections = [
        {
            'key': 'register',
            'title': '1️⃣ Зарегистрировать отчет',
            'icon': '📝'
        },
        {
            'key': 'automate', 
            'title': '2️⃣ Автоматизировать отчет',
            'icon': '🤖'
        },
        {
            'key': 'update',
            'title': '3️⃣ Актуализировать отчет', 
            'icon': '🔄'
        },
        {
            'key': 'change_owner',
            'title': '4️⃣ Сменить владельца отчета',
            'icon': '👤'
        },
        {
            'key': 'delete',
            'title': '5️⃣ Удалить отчет',
            'icon': '🗑️'
        }
    ]
    
    # Отображение разделов

    # Создаем вкладки
    tab_titles = [f"{section['icon']} {section['title']}" for section in sections]
    tabs = st.tabs(tab_titles)
    
    # Отображение содержимого каждой вкладки
    for i, (tab, section) in enumerate(zip(tabs, sections)):
        key = section['key']
        title = section['title']
        icon = section['icon']
        
        with tab:
            # Если админ режим включен, показываем редактор
            if st.session_state.admin_mode:
                st.markdown('<span class="admin-badge">РЕЖИМ РЕДАКТИРОВАНИЯ</span>', unsafe_allow_html=True)
                
                # Текстовое поле для редактирования
                edited_text = st.text_area(
                    f"Редактировать текст:",
                    value=st.session_state.action_texts[key],
                    height=200,
                    key=f"edit_{key}",
                    help="В админ режиме вы можете редактировать этот текст"
                )
                
                # Кнопки сохранения и сброса
                col1, col2, col3 = st.columns([1, 1, 4])
                
                with col1:
                    if st.button(f"💾 Сохранить", key=f"save_{key}", type="primary"):
                        st.session_state.action_texts[key] = edited_text
                        st.success("✅ Текст сохранен!")
                        st.rerun()
                
                with col2:
                    if st.button(f"🔄 Сбросить", key=f"reset_{key}"):
                        st.warning("⚠️ Сброс к исходному тексту")
                        st.rerun()
            
            else:
                # Обычный режим - просто показываем текст
                st.markdown(f"""
                <div style="
                    background-color: #f8f9fa;
                    padding: 1.5rem;
                    border-radius: 0.5rem;
                    border-left: 4px solid #1f77b4;
                    margin-bottom: 1rem;
                ">
                    <p style="margin: 0; line-height: 1.6;">{st.session_state.action_texts[key]}</p>
                </div>
                """, unsafe_allow_html=True)
    
    # Дополнительная информация внизу страницы
    st.markdown("### 📞 Нужна помощь?")
    st.info("""
    **Если у вас возникли вопросы по любому из действий:**
    - 📧 Напишите на email: support@company.com
    - 📞 Позвоните по телефону: +7 (xxx) xxx-xx-xx
    - 💬 Воспользуйтесь разделом "Задать вопрос (ИИ ассистент)"
    """)
    
    # Админ информация
    if st.session_state.admin_mode:
        st.markdown("---")
        st.markdown("### 🔧 Админ информация")
        st.markdown('<span class="admin-badge">АДМИН ПАНЕЛЬ</span>', unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        with col1:
            st.metric("Всего разделов", len(sections))
        with col2:
            # Подсчет общего количества символов во всех текстах
            total_chars = sum(len(text) for text in st.session_state.action_texts.values())
            st.metric("Общий объем текста", f"{total_chars} символов")
        
        # Кнопка экспорта всех текстов
        if st.button("📤 Экспортировать все тексты", key="export_all_texts"):
            export_data = {
                "exported_at": datetime.now().isoformat(),
                "sections": []
            }
            
            for section in sections:
                export_data["sections"].append({
                    "title": section["title"],
                    "key": section["key"],
                    "text": st.session_state.action_texts[section["key"]]
                })
            
            import json
            json_string = json.dumps(export_data, ensure_ascii=False, indent=2)
            
            st.download_button(
                label="💾 Скачать JSON с текстами",
                data=json_string,
                file_name=f"action_texts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json"
            )

def show_attributes():
    st.markdown('<div class="page-header">🏷️ Сформировать атрибуты и термины</div>', unsafe_allow_html=True)
    st.markdown("### Преобразование горизонтальной структуры Excel в вертикальные метаданные атрибутов")
    st.markdown("---")
    
    # Инициализация session state для атрибутов
    if 'attr_file_processed' not in st.session_state:
        st.session_state.attr_file_processed = False
    if 'attr_transformer' not in st.session_state:
        st.session_state.attr_transformer = None
    if 'attr_df' not in st.session_state:
        st.session_state.attr_df = None
    if 'attr_uploaded_file_name' not in st.session_state:
        st.session_state.attr_uploaded_file_name = ""
    
    # Основной интерфейс
    col1, col2 = st.columns([2, 1])
    
    with col2:
        st.header("📁 Загрузка файла")
        uploaded_file = st.file_uploader(
            "Выберите Excel или CSV файл",
            type=['xlsx', 'xls', 'csv'],
            help="Поддерживаются форматы: Excel (.xlsx, .xls) и CSV (.csv)",
            key="attributes_file_uploader"
        )
    
    with col1:
        st.header("⚙️ Параметры")
        report_number = st.text_input(
            "Номер отчета",
            value="R001",
            help="Этот номер будет использован для генерации кодов атрибутов и названия файла",
            key="attributes_report_number"
        )
        
        report_type = st.selectbox(
            "Тип отчета",
            options=["Ручной", "Полуавтоматический", "Автоматический", "ИЛА"],
            help="Тип отчета влияет на заполнение технических полей метаданных",
            key="attributes_report_type"
        )
    
    if uploaded_file is not None:
        st.session_state.attr_uploaded_file_name = uploaded_file.name
        
        # Создаем трансформер
        transformer = ExcelTransformer(report_number=report_number)
        
        try:
            # Загружаем данные
            df = transformer.load_from_uploaded_file(uploaded_file)
            
            # Сохраняем в session state
            st.session_state.attr_transformer = transformer
            st.session_state.attr_df = df
            st.session_state.attr_file_processed = True
            
            st.success("✅ Файл успешно загружен!")
            
            # Показываем базовую информацию о файле
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Количество строк", len(df))
            with col2:
                st.metric("Количество столбцов", len(df.columns))
            with col3:
                st.metric("Номер отчета", report_number)
            
            # Кнопка для выгрузки атрибутного состава
            st.markdown("---")
            if st.button("🔄 Выгрузить атрибутный состав", type="primary", use_container_width=True, key="generate_attributes"):
                with st.spinner("Преобразование данных..."):
                    # Сброс указателя файла и преобразование
                    uploaded_file.seek(0)
                    df_fresh = transformer.load_from_uploaded_file(uploaded_file)
                    metadata_df = transformer.transform_to_metadata(df_fresh, report_type)
                    
                    # Создаем Excel файл для скачивания
                    excel_data = transformer.create_excel_download(metadata_df)
                    
                    # Генерируем имя файла
                    filename = f"{report_number}_атрибуты.xlsx"
                    
                    st.success("✅ Атрибутный состав успешно создан!")
                    
                    # Показываем статистику
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Создано атрибутов", len(metadata_df))
                    with col2:
                        st.metric("Тип отчета", report_type)
                    with col3:
                        type_stats = metadata_df['base_type_report_field'].value_counts()
                        most_common_type = type_stats.index[0] if len(type_stats) > 0 else "N/A"
                        st.metric("Основной тип", most_common_type)
                    
                    # Кнопка скачивания
                    st.download_button(
                        label="📥 Скачать атрибутный состав",
                        data=excel_data,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                        key="download_attributes"
                    )
                    
                    st.info("💡 **Совет**: Файл необходимо доработать, заполнив обязательные колонки. А также проверить предзаполненные значения")
        
        except Exception as e:
            st.error(f"❌ Ошибка при обработке файла: {str(e)}")

def show_dashboard():
    st.markdown('<div class="page-header">📈 Дашборд по актуальности отчетов</div>', unsafe_allow_html=True)
    st.markdown('<div class="coming-soon">🚧 Здесь будут отображаться графики и метрики актуальности отчетов</div>', unsafe_allow_html=True)

def show_ai_assistant():
    st.markdown('<div class="page-header">🤖 Задать вопрос (ИИ ассистент)</div>', unsafe_allow_html=True)
    st.markdown('<div class="coming-soon">🚧 Здесь будет интеграция с ИИ-ассистентом для ответов на вопросы</div>', unsafe_allow_html=True)

def show_feedback():
    st.markdown('<div class="page-header">💬 Оставить обратную связь</div>', unsafe_allow_html=True)
    st.markdown('<div class="coming-soon">🚧 Здесь будет форма для отправки обратной связи и предложений</div>', unsafe_allow_html=True)

def show_admin_control():
    st.markdown('<div class="page-header">🔍 Контроль публикации отчетов</div>', unsafe_allow_html=True)
    st.markdown('<span class="admin-badge">АДМИН</span>', unsafe_allow_html=True)
    
    from utils import display_request_analysis
    
    st.markdown("## 📊 Анализатор запросов и стадий рассмотрения")
    st.info("Загрузите файл с запросами для анализа времени рассмотрения и текущих стадий")
    
    display_request_analysis()

def show_admin_stats():
    st.markdown('<div class="page-header">📊 Статистика по публикации</div>', unsafe_allow_html=True)
    st.markdown('<span class="admin-badge">АДМИН</span>', unsafe_allow_html=True)
    st.markdown('<div class="coming-soon">🚧 Административная статистика и аналитика по публикациям</div>', unsafe_allow_html=True)

def show_admin_issues():
    st.markdown('<div class="page-header">⚠️ Проблемные вопросы</div>', unsafe_allow_html=True)
    st.markdown('<span class="admin-badge">АДМИН</span>', unsafe_allow_html=True)
    st.markdown('<div class="coming-soon">🚧 Мониторинг и управление проблемными вопросами</div>', unsafe_allow_html=True)

# Отображение выбранной страницы
if st.session_state.admin_mode and 'selected_admin_page' in st.session_state:
    # Показываем админ страницы
    page_map = {
        "🔍 Контроль публикации отчетов": show_admin_control,
        "📊 Статистика по публикации": show_admin_stats, 
        "⚠️ Проблемные вопросы": show_admin_issues
    }
    page_map[st.session_state.selected_admin_page]()
else:
    # Показываем основные страницы
    page_map = {
        "📋 Инструкции по отчетам": show_instructions,
        "⚡ Действия с отчетами": show_actions,
        "🏷️ Сформировать атрибуты и термины": show_attributes,
        "📈 Дашборд по актуальности отчетов": show_dashboard,
        "🤖 Задать вопрос (ИИ ассистент)": show_ai_assistant,
        "💬 Оставить обратную связь": show_feedback
    }
    page_map[st.session_state.selected_page]()

# Подвал приложения
st.markdown("---")
col1, col2, col3 = st.columns(3)
with col1:
    st.caption("🏢 Компания XYZ")
with col2:
    st.caption("📧 support@company.com")
with col3:
    st.caption("📞 +7 (xxx) xxx-xx-xx")
