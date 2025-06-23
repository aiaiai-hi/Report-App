import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import io
from pathlib import Path
import warnings
from openpyxl import Workbook
from openpyxl.styles import Font
import os
import json

warnings.filterwarnings('ignore')

# Константы для хранения данных
DATA_DIR = Path("Дашборд")
REPORTS_DATA_FILE = DATA_DIR / "reports_data.xlsx"
COMMENTS_DATA_FILE = DATA_DIR / "comments_data.json"

# Добавьте новые константы для анализатора запросов
REQUESTS_DATA_FILE = DATA_DIR / "requests_data.xlsx"
REQUESTS_PROCESSED_FILE = DATA_DIR / "requests_processed.xlsx"

# Создаем директорию если её нет
DATA_DIR.mkdir(exist_ok=True)

# Конфигурация страницы
st.set_page_config(
    page_title="Система управления отчетами",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Инициализация состояния сессии
if 'admin_mode' not in st.session_state:
    st.session_state.admin_mode = False

# Функции для работы с данными
def save_reports_data(df, comments=None):
    """Сохранение данных отчетов в постоянный файл"""
    try:
        # Сохраняем основные данные
        df.to_excel(REPORTS_DATA_FILE, index=False)
        
        # Сохраняем комментарии
        if comments:
            with open(COMMENTS_DATA_FILE, 'w', encoding='utf-8') as f:
                json.dump(comments, f, ensure_ascii=False, indent=2)
        
        return True
    except Exception as e:
        st.error(f"Ошибка при сохранении данных: {str(e)}")
        return False

def save_requests_data(original_df, processed_df):
    """Сохранение данных анализатора запросов в постоянные файлы"""
    try:
        # Сохраняем исходные данные
        original_df.to_excel(REQUESTS_DATA_FILE, index=False)
        
        # Сохраняем обработанные данные
        processed_df.to_excel(REQUESTS_PROCESSED_FILE, index=False)
        
        return True
    except Exception as e:
        st.error(f"Ошибка при сохранении данных анализатора: {str(e)}")
        return False

def load_requests_data():
    """Загрузка данных анализатора запросов из постоянных файлов"""
    original_df = None
    processed_df = None
    
    try:
        # Загружаем исходные данные
        if REQUESTS_DATA_FILE.exists():
            original_df = pd.read_excel(REQUESTS_DATA_FILE)
        
        # Загружаем обработанные данные
        if REQUESTS_PROCESSED_FILE.exists():
            processed_df = pd.read_excel(REQUESTS_PROCESSED_FILE)
    
    except Exception as e:
        st.error(f"Ошибка при загрузке данных анализатора: {str(e)}")
    
    return original_df, processed_df

def init_requests_data():
    """Инициализация данных анализатора запросов при запуске приложения"""
    if 'requests_data_initialized' not in st.session_state:
        original_df, processed_df = load_requests_data()
        
        st.session_state.request_original_data = original_df
        st.session_state.request_processed_data = processed_df
        
        st.session_state.requests_data_initialized = True

def load_reports_data():
    """Загрузка данных отчетов из постоянного файла"""
    df = None
    comments = {}
    
    try:
        # Загружаем основные данные
        if REPORTS_DATA_FILE.exists():
            df = pd.read_excel(REPORTS_DATA_FILE)
        
        # Загружаем комментарии
        if COMMENTS_DATA_FILE.exists():
            with open(COMMENTS_DATA_FILE, 'r', encoding='utf-8') as f:
                comments_data = json.load(f)
                # Конвертируем строковые ключи обратно в int
                comments = {int(k): v for k, v in comments_data.items()}
    
    except Exception as e:
        st.error(f"Ошибка при загрузке данных: {str(e)}")
    
    return df, comments

def init_dashboard_data():
    """Инициализация данных дашборда при запуске приложения"""
    if 'reports_data_initialized' not in st.session_state:
        df, comments = load_reports_data()
        
        if df is not None:
            st.session_state.reports_data = df
            st.session_state.reports_comments = comments
        else:
            st.session_state.reports_data = None
            st.session_state.reports_comments = {}
        
        st.session_state.reports_data_initialized = True
    
    # Добавляем инициализацию данных анализатора
    init_requests_data()

# Инициализируем данные дашборда
init_dashboard_data()

# CSS стили для улучшения внешнего вида
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .page-header {
        font-size: 2rem;
        font-weight: bold;
        color: #333;
        border-bottom: 2px solid #1f77b4;
        padding-bottom: 0.5rem;
        margin-bottom: 1.5rem;
    }
    .sidebar-header {
        font-size: 1.2rem;
        font-weight: bold;
        color: #1f77b4;
        margin-bottom: 1rem;
    }
    .admin-badge {
        background-color: #ff4b4b;
        color: white;
        padding: 0.2rem 0.5rem;
        border-radius: 0.3rem;
        font-size: 0.8rem;
        font-weight: bold;
    }
    .coming-soon {
        text-align: center;
        color: #666;
        font-style: italic;
        font-size: 1.2rem;
        margin-top: 3rem;
        padding: 2rem;
        border: 2px dashed #ccc;
        border-radius: 0.5rem;
    }
    
    .section-container {
        background-color: #f8f9fa;
        padding: 1.5rem;
        border-radius: 0.5rem;
        border-left: 4px solid #1f77b4;
        margin-bottom: 1rem;
    }
    
    .edit-mode-badge {
        background-color: #ff8c00;
        color: white;
        padding: 0.2rem 0.5rem;
        border-radius: 0.3rem;
        font-size: 0.7rem;
        font-weight: bold;
        margin-bottom: 0.5rem;
        display: inline-block;
    }
    
    .section-text {
        margin: 0;
        line-height: 1.6;
        font-size: 1rem;
    }
    
    .stButton > button {
        width: 100% !important;
        background-color: transparent !important;
        border: none !important;
        border-radius: 0.5rem !important;
        padding: 0.4rem 0.75rem !important;
        text-align: left !important;
        color: #666 !important;
        font-weight: normal !important;
        margin-bottom: 0 !important;
        transition: all 0.2s ease !important;
        justify-content: flex-start !important;
        display: flex !important;
        align-items: center !important;
        min-height: 2.5rem !important;
        line-height: 1.1 !important;
    }
    
    .stButton {
        margin-bottom: 0rem !important;
    }
    
    [data-testid="stSidebar"] .stButton {
        margin-bottom: 0rem !important;
    }
 
    [data-testid="stSidebar"] .element-container {
        margin-bottom: 0rem !important;
    }        

    .stButton > button:hover {
        background-color: #d1d3d4 !important;
        color: #333 !important;
    }
    
    .stButton > button:focus {
        background-color: #d1d3d4 !important;
        color: #333 !important;
        box-shadow: none !important;
        outline: none !important;
    }
    
    .stButton > button[data-testid*="baseButton-secondary"] {
        background-color: #d1d3d4 !important;
        color: #333 !important;
        font-weight: 500 !important;
    }
    
    .stButton > button[kind="secondary"] {
        background-color: #d1d3d4 !important;
        color: #333 !important;
        font-weight: 500 !important;
    }
</style>
""", unsafe_allow_html=True)

# Класс ExcelTransformer
class ExcelTransformer:
    def __init__(self, report_number=None):
        """
        Инициализация трансформера
        
        Args:
            report_number (str): Номер отчета для генерации кодов атрибутов
        """
        self.report_number = report_number or "R001"
        self.supported_extensions = ['.xlsx', '.xls', '.csv']
        self.report_types = ["Ручной", "Полуавтоматический", "Автоматический", "ИЛА"]
    
    def detect_data_type(self, values):
        """
        Автоматическое определение типа данных столбца
        
        Args:
            values: pandas Series с данными столбца
            
        Returns:
            str: тип данных ('текст', 'число', 'дата', 'флаг')
        """
        # Убираем пустые значения и NaN
        clean_values = values.dropna()
        if len(clean_values) == 0:
            return "текст"
        
        # Конвертируем в строки для анализа
        str_values = clean_values.astype(str).str.strip().str.lower()
        
        # Проверка на булевы значения (флаги)
        bool_indicators = {
            'да', 'нет', 'true', 'false', '1', '0', 'yes', 'no', 
            'y', 'n', 'вкл', 'выкл', 'on', 'off', 'активен', 'неактивен'
        }
        unique_values = set(str_values.unique())
        if unique_values.issubset(bool_indicators) and len(unique_values) <= 3:
            return "флаг"
        
        # Проверка на даты
        date_count = 0
        for val in clean_values:
            if self._is_date(val):
                date_count += 1
        
        if date_count / len(clean_values) > 0.7:  # 70% значений - даты
            return "дата"
        
        # Проверка на числа
        numeric_count = 0
        for val in clean_values:
            if self._is_numeric(val):
                numeric_count += 1
        
        if numeric_count / len(clean_values) > 0.8:  # 80% значений - числа
            return "число"
        
        return "текст"
    
    def _is_date(self, value):
        """Проверка, является ли значение датой"""
        if pd.isna(value):
            return False
            
        # Попробуем распарсить как дату
        date_formats = [
            '%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y',
            '%d.%m.%y', '%d/%m/%y', '%y-%m-%d', '%d-%m-%y',
            '%Y.%m.%d', '%Y/%m/%d'
        ]
        
        str_val = str(value).strip()
        
        for fmt in date_formats:
            try:
                datetime.strptime(str_val, fmt)
                return True
            except ValueError:
                continue
                
        # Проверим pandas to_datetime
        try:
            pd.to_datetime(str_val, errors='raise')
            return True
        except:
            return False
    
    def _is_numeric(self, value):
        """Проверка, является ли значение числом"""
        if pd.isna(value):
            return False
            
        try:
            # Попробуем конвертировать в float
            float(str(value).replace(',', '.').replace(' ', ''))
            return True
        except ValueError:
            return False
    
    def load_from_uploaded_file(self, uploaded_file):
        """
        Загрузка данных из uploaded_file Streamlit
        
        Args:
            uploaded_file: файл, загруженный через st.file_uploader
            
        Returns:
            pandas.DataFrame: загруженные данные
        """
        try:
            file_extension = Path(uploaded_file.name).suffix.lower()
            
            if file_extension == '.csv':
                # Для CSV пробуем разные разделители и кодировки
                try:
                    df = pd.read_csv(uploaded_file, sep=',', encoding='utf-8')
                    if len(df.columns) > 1:
                        return df
                except:
                    pass
                
                try:
                    uploaded_file.seek(0)  # Сброс указателя файла
                    df = pd.read_csv(uploaded_file, sep=';', encoding='cp1251')
                    if len(df.columns) > 1:
                        return df
                except:
                    pass
                
                # Последняя попытка
                uploaded_file.seek(0)
                df = pd.read_csv(uploaded_file)
                
            else:
                # Для Excel файлов
                df = pd.read_excel(uploaded_file)
            
            return df
            
        except Exception as e:
            raise Exception(f"Ошибка при загрузке файла: {str(e)}")
    
    def transform_to_metadata(self, df, report_type):
        """
        Преобразование DataFrame в метаданные атрибутов
        
        Args:
            df (pandas.DataFrame): исходные данные
            report_type (str): тип отчета (Ручной, Полуавтоматический, Автоматический, ИЛА)
            
        Returns:
            pandas.DataFrame: метаданные атрибутов
        """
        metadata_list = []
        
        for idx, column in enumerate(df.columns, 1):
            # Получаем данные столбца
            column_data = df[column]
            
            # Определяем тип данных
            data_type = self.detect_data_type(column_data)
            
            # Определяем значения по умолчанию в зависимости от типа отчета
            if report_type in ["Ручной", "Полуавтоматический"]:
                tech_algorithm_to_be = "Ручной ввод"
                data_source_type = "Ручное заполнение"
            else:  # Автоматический или ИЛА
                tech_algorithm_to_be = ""
                data_source_type = "База данных"
            
            # Связь с ИС для ИЛА
            system_connection = "ИЛА One" if report_type == "ИЛА" else ""
            
            metadata_record = {
                'ReportCode_info': '',  # Будет заполнено позже
                'Noreportfield_info': idx,
                'name': column,
                'description': '',
                'TechAsIs': '',
                'BussAlgorythm': '',
                'TechAlgorythm': tech_algorithm_to_be,
                'algorithms_change_info': 'нет',
                'dbobjectlink': '',
                'base_type_info': data_source_type,
                'related_it_system_info': system_connection,
                'reportfields_codes': '',
                'reportfields_names': '',
                'reportfields_parent_term': '',
                'reportfields_domain': '',
                'required_attribute_info': 'да',
                'base_type_report_field': data_type,
                'base_calc_ref_ind_info': 'Базовый',
                'codeTable_info': '',
                'example': '',
                'isToDelete_info': ''
            }
            
            metadata_list.append(metadata_record)
        
        metadata_df = pd.DataFrame(metadata_list)
        
        # Заполняем код атрибута после создания DataFrame
        metadata_df['ReportCode_info'] = metadata_df['Noreportfield_info'].apply(
            lambda x: f"{self.report_number}_{x:03d}"
        )
        
        return metadata_df
    
    def create_excel_download(self, metadata_df):
        """
        Создание Excel файла для скачивания с заголовками и пользовательскими названиями
        
        Args:
            metadata_df (pandas.DataFrame): метаданные для сохранения
            
        Returns:
            bytes: данные Excel файла
        """
        output = io.BytesIO()
        
        # Создаем новую книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Атрибут отчета"
        
        # Технические заголовки (скрытая строка)
        technical_headers = [
            'ReportCode_info', 'Noreportfield_info', 'name', 'description', 'TechAsIs', 
            'BussAlgorythm', 'TechAlgorythm', 'algorithms_change_info', 'dbobjectlink', 
            'base_type_info', 'related_it_system_info', 'reportfields_codes', 
            'reportfields_names', 'reportfields_parent_term', 'reportfields_domain', 
            'required_attribute_info', 'base_type_report_field', 'base_calc_ref_ind_info', 
            'codeTable_info', 'example', 'isToDelete_info'
        ]
        
        # Пользовательские заголовки (видимая строка)
        user_headers = [
            'Код атрибута отчета', 
            '№ атрибута отчета', 
            'Наименование атрибута', 
            'Бизнес-алгоритм AS IS',
            'Технический алгоритм AS IS', 
            'Бизнес-алгоритм TO BE',
            'Технический алгоритм TO BE', 
            'Алгоритм изменен', 
            'Физические атрибуты', 
            'Тип источника данных',
            'Связь с информационной системой', 
            'Код термина/терминов',
            'Наименование термина/терминов', 
            'Наименование родительской сущности термина/терминов', 
            'Домен термина/терминов', 
            'Обязательный атрибут для заполнения', 
            'Базовый тип атрибута (Текст, Число, Дата, Флаг)', 
            'Признак атрибута (Базовый, Расчетный, Справочный)', 
            'Наименование справочника', 
            'Примечание', 
            'Помечен к удалению (да/нет)'
        ]
        
        # Записываем технические заголовки в первую строку (скрытую)
        for col_idx, header in enumerate(technical_headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Скрываем первую строку
        ws.row_dimensions[1].hidden = True
        
        # Записываем пользовательские заголовки во вторую строку (видимую)
        for col_idx, header in enumerate(user_headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            # Делаем заголовки полужирными
            cell.font = Font(bold=True)
        
        # Закрепляем первые две строки
        ws.freeze_panes = ws.cell(row=3, column=1)
        
        # Записываем данные начиная с третьей строки
        for row_idx, (_, row) in enumerate(metadata_df.iterrows(), 3):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Автоподбор ширины столбцов (учитываем все строки включая заголовки)
        for col_idx, column_letter in enumerate([chr(65 + i) for i in range(len(technical_headers))], 0):
            max_length = 0
            
            # Проверяем техническую строку
            if len(technical_headers) > col_idx:
                if len(str(technical_headers[col_idx])) > max_length:
                    max_length = len(str(technical_headers[col_idx]))
            
            # Проверяем пользовательскую строку
            if len(user_headers) > col_idx:
                if len(str(user_headers[col_idx])) > max_length:
                    max_length = len(str(user_headers[col_idx]))
            
            # Проверяем данные
            for row_idx in range(3, len(metadata_df) + 3):
                cell_value = ws.cell(row=row_idx, column=col_idx + 1).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в BytesIO
        wb.save(output)
        output.seek(0)
        return output.getvalue()

# Функции для дашборда
def calculate_completion_percentage(df, owner_filter=None):
    """Расчет процента заполнения полей"""
    if df is None or df.empty:
        return 0, 0
    
    # Фильтруем по владельцу если задан
    if owner_filter and owner_filter != "Все":
        df = df[df['Владелец отчета ССП'] == owner_filter]
    
    if df.empty:
        return 0, 0
    
    total_cells = 0
    filled_cells = 0
    
    for _, row in df.iterrows():
        for col in df.columns:
            # Исключения для расчета
            participation_value = row.get('Участие в формировании РФ', '')
            if col == 'ССП, в функциональном подчинении которого, находятся сотрудники РФ' and str(participation_value).lower() == 'нет':
                continue
            frequency_value = row.get('Частота отчета', '')
            if col == 'Частота отчета (ручной ввод)' and str(frequency_value).lower() != 'ручной ввод':
                continue
            
            total_cells += 1
            if pd.notna(row[col]) and str(row[col]).strip() != '':
                filled_cells += 1
    
    completion_rate = (filled_cells / total_cells * 100) if total_cells > 0 else 0
    
    # Процент опубликованных отчетов
    published_count = len(df[df['Этап отчета'] == 'Опубликован'])
    total_reports = len(df)
    published_rate = (published_count / total_reports * 100) if total_reports > 0 else 0
    
    return completion_rate, published_rate

def get_reports_needing_confirmation(df):
    """Отчеты, требующие подтверждения актуальности"""
    if df is None or df.empty:
        return pd.DataFrame()
    
    from datetime import datetime, timedelta
    import pandas as pd
    
    result_data = []
    current_date = datetime.now()
    
    for _, row in df.iterrows():
        last_publication = row.get('Дата последней публикации отчета')
        if pd.notna(last_publication):
            try:
                if isinstance(last_publication, str):
                    pub_date = pd.to_datetime(last_publication)
                else:
                    pub_date = last_publication
                
                actualization_date = pub_date + timedelta(days=365)  # +1 год
                days_until_actualization = (actualization_date - current_date).days
                
                if days_until_actualization <= 60:  # 2 месяца или менее
                    # Форматируем дату в дд.мм.гггг
                    pub_date_formatted = pub_date.strftime('%d.%m.%Y')
                    actualization_date_formatted = actualization_date.strftime('%d.%m.%Y')
                    
                    # Рассчитываем месяцы и дни
                    if days_until_actualization < 0:
                        # Просрочено
                        abs_days = abs(days_until_actualization)
                        months = abs_days // 30
                        days = abs_days % 30
                        status_text = f"Просрочено {months} месяцев, {days} дней"
                        status_color = "🔴"
                    else:
                        # Не просрочено
                        months = days_until_actualization // 30
                        days = days_until_actualization % 30
                        status_text = f"Осталось {months} месяцев, {days} дней"
                        status_color = "🟢"
                    
                    result_data.append({
                        'Номер формы': row.get('Номер формы', ''),
                        'Наименование отчета': row.get('Наименование отчета', ''),
                        'Владелец отчета ССП': row.get('Владелец отчета ССП', ''),
                        'Дата последней публикации': pub_date_formatted,
                        'Дата актуализации': actualization_date_formatted,
                        'Статус актуализации': f"{status_color} {status_text}"
                    })
            except:
                continue
    
    return pd.DataFrame(result_data)

def get_reports_needing_update(df):
    """Отчеты, требующие актуализации"""
    if df is None or df.empty:
        return pd.DataFrame()
    
    result_data = []
    
    for _, row in df.iterrows():
        needs_update = False
        actions = []
        comments = []
        
        # Проверка статуса
        status = row.get('Этап отчета', '')
        if status != 'Опубликован':
            needs_update = True
            actions.append("Необходимо довести отчет до публикации")
        
        # Проверка незаполненных полей
        empty_fields = []
        for col in df.columns:
            # Исключения
            participation_value = row.get('Участие в формировании РФ', '')
            if col == 'ССП, в функциональном подчинении которого, находятся сотрудники РФ' and str(participation_value).lower() == 'нет':
                continue
            frequency_value = row.get('Частота отчета', '')
            if col == 'Частота отчета (ручной ввод)' and str(frequency_value).lower() != 'ручной ввод':
                continue
            
            if pd.isna(row[col]) or str(row[col]).strip() == '':
                empty_fields.append(col)
        
        if empty_fields:
            needs_update = True
            if status == 'Опубликован':
                actions.append("Создать запрос на актуализацию")
            comments.append(f"Заполнить поля ({'; '.join(empty_fields)})")
        
        # Проверка шаблона
        template_value = row.get('Шаблон отчета', '')
        if str(template_value).lower() == 'нет':
            needs_update = True
            comments.append("Добавить шаблон")
        
        # Проверка атрибутов
        attributes_value = row.get('Атрибуты описаны', '')
        if str(attributes_value).lower() == 'нет':
            needs_update = True
            comments.append("Описать атрибуты")
        
        if needs_update:
            result_data.append({
                'Номер формы': row.get('Номер формы', ''),
                'Наименование отчета': row.get('Наименование отчета', ''),
                'Этап отчета': status,
                'Владелец отчета ССП': row.get('Владелец отчета ССП', ''),
                'Необходимые действия': '; '.join(actions),
                'Доп. комментарии': '; '.join(comments)
            })
    
    return pd.DataFrame(result_data)

# Заголовок приложения
st.markdown('<div class="main-header">📊 Система управления отчетами</div>', unsafe_allow_html=True)

# Боковая панель с навигацией
with st.sidebar:
    st.markdown('<div class="sidebar-header">🧭 Навигация</div>', unsafe_allow_html=True)
    
    # Основное меню
    st.markdown("**Основные функции:**")
    main_pages = {
        "📈 Дашборд по отчетам": "dashboard",
        "⚡ Действия с отчетами": "actions",
        "🏷️ Сформировать атрибуты": "attributes", 
        "❓ Частые вопросы": "ai_assistant",
        "📋 Документация": "instructions",
        "💬 Оставить обратную связь": "feedback"
    }
    
    # Для основного меню:
    if 'selected_page' not in st.session_state:
        st.session_state.selected_page = "📈 Дашборд по отчетам"
    
    for page_name in main_pages.keys():
        # Определяем, является ли эта кнопка активной
        is_active = st.session_state.selected_page == page_name
        button_type = "secondary" if is_active else "primary"
        
        if st.button(page_name, key=f"btn_{main_pages[page_name]}", use_container_width=True, type=button_type):
            st.session_state.selected_page = page_name
            st.rerun()
    
    st.markdown("---")
    
    # Админ панель
    st.markdown("**Административная панель:**")
    
    # Инициализация состояния авторизации
    if 'admin_authenticated' not in st.session_state:
        st.session_state.admin_authenticated = False
    
    if not st.session_state.admin_authenticated:
        # Форма авторизации
        st.markdown("🔐 **Админ вход**")
        
        with st.form("admin_login_form"):
            username = st.text_input("Имя пользователя:", placeholder="admin")
            password = st.text_input("Пароль:", type="password", placeholder="Введите пароль")
            login_button = st.form_submit_button("🔑 Войти", use_container_width=True)
            
            if login_button:
                if username == "admin" and password == "!!!!QQQQ2222":
                    st.session_state.admin_authenticated = True
                    st.session_state.admin_mode = True
                    st.success("✅ Успешная авторизация!")
                    st.rerun()
                else:
                    st.error("❌ Неверное имя пользователя или пароль!")
    else:
        # Админ уже авторизован
        st.session_state.admin_mode = True
        st.markdown('<span class="admin-badge">АДМИН РЕЖИМ</span>', unsafe_allow_html=True)
        
        if st.button("🚪 Выйти", key="admin_logout", use_container_width=True):
            st.session_state.admin_authenticated = False
            st.session_state.admin_mode = False
            st.success("👋 Вы вышли из админ режима")
            st.rerun()
        st.markdown("")
    
    if st.session_state.admin_mode:
        
        admin_pages = {
            "🔍 Контроль публикации отчетов": "admin_control",
            "📊 Дашборд по отчетам": "admin_dashboard",
            "📈 Статистика по публикации": "admin_stats", 
            "⚠️ Проблемные вопросы": "admin_issues"
        }
        
        # Для админ меню:
        if 'selected_admin_page' not in st.session_state:
            st.session_state.selected_admin_page = None  # Не выбираем админ страницу по умолчанию
            
        for page_name in admin_pages.keys():
            # Определяем, является ли эта админ кнопка активной
            is_active = st.session_state.selected_admin_page == page_name
            button_type = "secondary" if is_active else "primary"
            
            if st.button(page_name, key=f"admin_btn_{admin_pages[page_name]}", use_container_width=True, type=button_type):
                st.session_state.selected_admin_page = page_name
                st.rerun()

# Функции для отображения страниц
def show_instructions():
    st.markdown('<div class="page-header">📋 Документация</div>', unsafe_allow_html=True)
    st.markdown('<div class="coming-soon">🚧 Здесь будут размещены подробные инструкции по работе с отчетами</div>', unsafe_allow_html=True)

def show_actions():
    st.markdown('<div class="page-header">⚡ Действия с отчетами</div>', unsafe_allow_html=True)
    
    # Инициализация состояния для редактируемых текстов
    if 'action_texts' not in st.session_state:
        st.session_state.action_texts = {
            'register': 'Чтобы зарегистрировать новый отчет необходимо зайти в Бизнес-глоссарий и в левом меню найти раздел "Запросы - Отчеты". В правом верхнем углу выбрать кнопку "Создать".',
            'automate': 'Если у вас уже существует отчет, который собирается регулярно "вручную", проверьте наличие зарегистрированной информации об этом отчете в Бизнес-глоссарии. Если "ручной отчет" отсутствует в реестре отчетов, то прежде, чем направлять заявку на автоматизацию, необходимо пройти регистрацию "ручного отчета". После этого можете переходить к заявке на автоматизацию отчета. Автоматизация существующего ручного отчета/ автоматизация нового отчета ("ручной" отчет по аналогичной форме не собирается с ССП или РФ). Для подачи заявки на автоматизацию отчета, необходимо сформировать запрос в Бизнес-глоссарии: в левом меню найти раздел "Запросы - Отчеты". В правом верхнем углу выбрать кнопку "Создать".',
            'update': 'Чтобы актуализировать существующий отчет ...',
            'change_owner': 'Чтобы сменить владельца отчета, необходимо выбрать 1 из 2х вариантов: 1. Направить служебную записку в адрес ДБД в свободной форме или 2. Создать запрос на смену владельца отчета в Бизнес-глоссарии, предварительно уточнив ФИО ответственного, кто будет принимать отчет, чтобы указать его владельцем запроса после передачи отчета новому владельцу',
            'delete': 'Чтобы удалить отчет, необходимо сформировать запрос на удаление отчета. Если отчет автоматизирован, то необходимо приложить BIQ, уточнить ответственного за автоматизацию, чтобы передать данную информацию для отключения отчета в системе.'
        }
    
    # Разделы с действиями
    sections = [
        {
            'key': 'register',
            'title': '1️⃣ Зарегистрировать отчет',
            'icon': '📝'
        },
        {
            'key': 'automate', 
            'title': '2️⃣ Автоматизировать отчет',
            'icon': '🤖'
        },
        {
            'key': 'update',
            'title': '3️⃣ Актуализировать отчет', 
            'icon': '🔄'
        },
        {
            'key': 'change_owner',
            'title': '4️⃣ Сменить владельца отчета',
            'icon': '👤'
        },
        {
            'key': 'delete',
            'title': '5️⃣ Удалить отчет',
            'icon': '🗑️'
        }
    ]
    
    # Создаем вкладки
    tab_titles = [f"{section['icon']} {section['title']}" for section in sections]
    tabs = st.tabs(tab_titles)
    
    # Отображение содержимого каждой вкладки
    for i, (tab, section) in enumerate(zip(tabs, sections)):
        key = section['key']
        title = section['title']
        icon = section['icon']
        
        with tab:
            # Если админ режим включен, показываем редактор
            if st.session_state.admin_mode:
                st.markdown('<span class="admin-badge">РЕЖИМ РЕДАКТИРОВАНИЯ</span>', unsafe_allow_html=True)
                
                # Текстовое поле для редактирования
                edited_text = st.text_area(
                    f"Редактировать текст:",
                    value=st.session_state.action_texts[key],
                    height=200,
                    key=f"edit_{key}",
                    help="В админ режиме вы можете редактировать этот текст"
                )
                
                # Кнопки сохранения и сброса
                col1, col2, col3 = st.columns([1, 1, 4])
                
                with col1:
                    if st.button(f"💾 Сохранить", key=f"save_{key}", type="primary"):
                        st.session_state.action_texts[key] = edited_text
                        st.success("✅ Текст сохранен!")
                        st.rerun()
                
                with col2:
                    if st.button(f"🔄 Сбросить", key=f"reset_{key}"):
                        st.warning("⚠️ Сброс к исходному тексту")
                        st.rerun()
            
            else:
                # Обычный режим - просто показываем текст
                st.markdown(f"""
                <div style="
                    background-color: #f8f9fa;
                    padding: 1.5rem;
                    border-radius: 0.5rem;
                    border-left: 4px solid #1f77b4;
                    margin-bottom: 1rem;
                ">
                    <p style="margin: 0; line-height: 1.6;">{st.session_state.action_texts[key]}</p>
                </div>
                """, unsafe_allow_html=True)
    
    # Дополнительная информация внизу страницы
    st.markdown("### 📞 Нужна помощь?")
    st.info("""
    **Если у вас возникли вопросы по любому из действий:**
    - 📧 Напишите на email: support@company.com
    - 📞 Позвоните по телефону: +7 (xxx) xxx-xx-xx
    - 💬 Воспользуйтесь разделом "Задать вопрос (ИИ ассистент)"
    """)
    
    # Админ информация
    if st.session_state.admin_mode:
        st.markdown("---")
        st.markdown("### 🔧 Админ информация")
        st.markdown('<span class="admin-badge">АДМИН ПАНЕЛЬ</span>', unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        with col1:
            st.metric("Всего разделов", len(sections))
        with col2:
            # Подсчет общего количества символов во всех текстах
            total_chars = sum(len(text) for text in st.session_state.action_texts.values())
            st.metric("Общий объем текста", f"{total_chars} символов")
        
        # Кнопка экспорта всех текстов
        if st.button("📤 Экспортировать все тексты", key="export_all_texts"):
            export_data = {
                "exported_at": datetime.now().isoformat(),
                "sections": []
            }
            
            for section in sections:
                export_data["sections"].append({
                    "title": section["title"],
                    "key": section["key"],
                    "text": st.session_state.action_texts[section["key"]]
                })
            
            import json
            json_string = json.dumps(export_data, ensure_ascii=False, indent=2)
            
            st.download_button(
                label="💾 Скачать JSON с текстами",
                data=json_string,
                file_name=f"action_texts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json"
            )

def show_attributes():
    st.markdown('<div class="page-header">🏷️ Сформировать атрибуты</div>', unsafe_allow_html=True)
    st.markdown("### Преобразование горизонтальной структуры Excel в вертикальные метаданные атрибутов")
    st.markdown("---")
    
    # Инициализация session state для атрибутов
    if 'attr_file_processed' not in st.session_state:
        st.session_state.attr_file_processed = False
    if 'attr_transformer' not in st.session_state:
        st.session_state.attr_transformer = None
    if 'attr_df' not in st.session_state:
        st.session_state.attr_df = None
    if 'attr_uploaded_file_name' not in st.session_state:
        st.session_state.attr_uploaded_file_name = ""
    
    # Основной интерфейс
    col1, col2 = st.columns([1, 1])
    
    with col2:
        st.header("📁 Загрузка файла")
        uploaded_file = st.file_uploader(
            "Выберите Excel или CSV файл",
            type=['xlsx', 'xls', 'csv'],
            help="Поддерживаются форматы: Excel (.xlsx, .xls) и CSV (.csv)",
            key="attributes_file_uploader"
        )
    
    with col1:
        st.header("⚙️ Параметры")
        report_number = st.text_input(
            "Номер отчета",
            value="R001",
            help="Этот номер будет использован для генерации кодов атрибутов и названия файла",
            key="attributes_report_number"
        )
        
        report_type = st.selectbox(
            "Тип отчета",
            options=["Ручной", "Полуавтоматический", "Автоматический", "ИЛА"],
            help="Тип отчета влияет на заполнение технических полей метаданных",
            key="attributes_report_type"
        )
    
    if uploaded_file is not None:
        st.session_state.attr_uploaded_file_name = uploaded_file.name
        
        # Создаем трансформер
        transformer = ExcelTransformer(report_number=report_number)
        
        try:
            # Загружаем данные
            df = transformer.load_from_uploaded_file(uploaded_file)
            
            # Сохраняем в session state
            st.session_state.attr_transformer = transformer
            st.session_state.attr_df = df
            st.session_state.attr_file_processed = True
            
            st.success("✅ Файл успешно загружен!")
            
            # Показываем базовую информацию о файле
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Количество строк", len(df))
            with col2:
                st.metric("Количество столбцов", len(df.columns))
            with col3:
                st.metric("Номер отчета", report_number)
            
            # Кнопка для выгрузки атрибутного состава
            st.markdown("---")
            if st.button("🔄 Выгрузить атрибутный состав", type="primary", use_container_width=True, key="generate_attributes"):
                with st.spinner("Преобразование данных..."):
                    # Сброс указателя файла и преобразование
                    uploaded_file.seek(0)
                    df_fresh = transformer.load_from_uploaded_file(uploaded_file)
                    metadata_df = transformer.transform_to_metadata(df_fresh, report_type)
                    
                    # Создаем Excel файл для скачивания
                    excel_data = transformer.create_excel_download(metadata_df)
                    
                    # Генерируем имя файла
                    filename = f"{report_number}_атрибуты.xlsx"
                    
                    st.success("✅ Атрибутный состав успешно создан!")
                    
                    # Показываем статистику
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Создано атрибутов", len(metadata_df))
                    with col2:
                        st.metric("Тип отчета", report_type)
                    with col3:
                        type_stats = metadata_df['base_type_report_field'].value_counts()
                        most_common_type = type_stats.index[0] if len(type_stats) > 0 else "N/A"
                        st.metric("Основной тип", most_common_type)
                    
                    # Кнопка скачивания
                    st.download_button(
                        label="📥 Скачать атрибутный состав",
                        data=excel_data,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                        key="download_attributes"
                    )
                    
                    st.info("💡 **Совет**: Файл необходимо доработать, заполнив обязательные колонки. А также проверить предзаполненные значения")
        
        except Exception as e:
            st.error(f"❌ Ошибка при обработке файла: {str(e)}")

def show_dashboard():
    st.markdown('<div class="page-header">📈 Дашборд по отчетам</div>', unsafe_allow_html=True)
    
    # Проверяем наличие данных (теперь данные доступны всегда)
    if st.session_state.reports_data is None:
        st.warning("⚠️ Данные не загружены.")
        
        # Кнопка для принудительной загрузки данных
        if st.button("🔄 Попробовать загрузить данные из файла"):
            df, comments = load_reports_data()
            if df is not None:
                st.session_state.reports_data = df
                st.session_state.reports_comments = comments
                st.success("✅ Данные успешно загружены!")
                st.rerun()
            else:
                st.error("❌ Файл данных не найден. Загрузите данные через Админ панель.")
        
        return
    
    # Показываем информацию о данных
    df = st.session_state.reports_data.copy()
    
    # Информационная панель
    #with st.expander("📋 Информация о данных", expanded=False):
    #    col1, col2, col3 = st.columns(3)
    #    with col1:
    #        st.metric("Всего отчетов", len(df))
    #    with col2:
    #        comments_count = len([c for c in st.session_state.reports_comments.values() if c.strip()])
    #        st.metric("Комментариев", comments_count)
    #    with col3:
    #        if REPORTS_DATA_FILE.exists():
    #            file_time = datetime.fromtimestamp(REPORTS_DATA_FILE.stat().st_mtime)
    #            st.metric("Обновлено", file_time.strftime('%d.%m.%Y'))

    # Фильтр по владельцу ССП
    st.markdown("## 🎯 Фильтры")
    if 'Владелец отчета ССП' in df.columns:
        owners = ['Все'] + sorted(df['Владелец отчета ССП'].dropna().unique().tolist())
        selected_owner = st.selectbox("Выберите владельца отчета ССП", owners, key="dashboard_owner_filter")
    else:
        selected_owner = "Все"
    
    # Применяем фильтр
    filtered_df = df.copy()
    if selected_owner != "Все":
        filtered_df = filtered_df[filtered_df['Владелец отчета ССП'] == selected_owner]
    
    # Основные метрики
    st.markdown("## 📊 Ключевые показатели")
    completion_rate, published_rate = calculate_completion_percentage(filtered_df)
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric(
            "% заполнения полей",
            f"{completion_rate:.1f}%",
            delta=None,
            help="Процент заполненных полей в отчетах"
        )
    
    with col2:
        st.metric(
            "% опубликованных отчетов",
            f"{published_rate:.1f}%",
            delta=None,
            help="Процент отчетов в статусе 'Опубликован'"
        )
    
    with col3:
        st.metric(
            "Всего отчетов",
            len(filtered_df),
            delta=None,
            help="Общее количество отчетов"
        )
    
    # Рекомендации
    st.markdown("---")
    st.markdown("## 💡 Рекомендации по отчетам")
    
    # 1. Отчеты, требующие подтверждения актуальности
    st.markdown("### 🔔 Необходимо подтверждение актуальности отчетов")
    confirmation_reports = get_reports_needing_confirmation(filtered_df)
    
    if not confirmation_reports.empty:
        st.dataframe(
            confirmation_reports,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Номер формы": st.column_config.TextColumn("Номер формы", width="small"),
                "Наименование отчета": st.column_config.TextColumn("Наименование отчета", width="medium"),
                "Владелец отчета ССП": st.column_config.TextColumn("Владелец отчета ССП", width="small"),
                "Дата последней публикации": st.column_config.TextColumn("Дата последней публикации", width="small"),
                "Дата актуализации": st.column_config.TextColumn("Дата актуализации", width="small"),
                "Статус актуализации": st.column_config.TextColumn("Статус актуализации", width="small")
            }
        )
    else:
        st.info("✅ Нет отчетов, требующих подтверждения актуальности в ближайшие 2 месяца")
    
    # 2. Отчеты, требующие актуализации
    st.markdown("### ⚠️ Требуется актуализация отчетов")
    update_reports = get_reports_needing_update(filtered_df)
    
    if not update_reports.empty:
        st.dataframe(
            update_reports,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Номер формы": st.column_config.TextColumn("Номер формы", width="small"),
                "Наименование отчета": st.column_config.TextColumn("Наименование отчета", width="medium"),
                "Этап отчета": st.column_config.TextColumn("Этап отчета", width="small"),
                "Владелец отчета ССП": st.column_config.TextColumn("Владелец отчета ССП", width="small"),
                "Необходимые действия": st.column_config.TextColumn("Необходимые действия", width="small"),
                "Доп. комментарии": st.column_config.TextColumn("Доп. комментарии", width="smallm")
            }
        )
    else:
        st.info("✅ Все отчеты актуальны")
    
    # Детальная таблица с фильтрами
    st.markdown("---")
    st.markdown("## 📋 Детальная информация по отчетам")
    
    # Дополнительные фильтры
    col1, col2 = st.columns(2)
    with col1:
        if 'Этап отчета' in df.columns:
            stages = ['Все'] + sorted(filtered_df['Этап отчета'].dropna().unique().tolist())
            selected_stage = st.selectbox("Фильтр по этапу отчета", stages, key="dashboard_stage_filter")
        else:
            selected_stage = "Все"
    
    with col2:
        if 'Тип формирования отчета' in df.columns:
            types = ['Все'] + sorted(filtered_df['Тип формирования отчета'].dropna().unique().tolist())
            selected_type = st.selectbox("Фильтр по типу формирования", types, key="dashboard_type_filter")
        else:
            selected_type = "Все"
    
    # Применяем дополнительные фильтры
    if selected_stage != "Все":
        filtered_df = filtered_df[filtered_df['Этап отчета'] == selected_stage]
    if selected_type != "Все":
        filtered_df = filtered_df[filtered_df['Тип формирования отчета'] == selected_type]
    
    # Функция для стилизации пустых ячеек
    def highlight_empty_cells(df):
        def style_func(val):
            if pd.isna(val) or str(val).strip() == '':
                return 'background-color: #ffcccc'
            return ''
        return df.style.applymap(style_func)
    
    # Форматируем даты в основной таблице
    display_df = filtered_df.copy()
    
    # Форматируем даты в дд.мм.гггг
    date_columns = ['Дата создания последнего черновика', 'Дата последней публикации отчета']
    for col in date_columns:
        if col in display_df.columns:
            display_df[col] = pd.to_datetime(display_df[col], errors='coerce').dt.strftime('%d.%m.%Y')
            display_df[col] = display_df[col].replace('NaT', '')
    
    # Создаем конфигурацию для всех столбцов с одинаковой шириной
    column_config = {}
    for col in display_df.columns:
        column_config[col] = st.column_config.TextColumn(col, width="small")
    
    # Отображаем стилизованную таблицу
    styled_df = highlight_empty_cells(display_df)
    st.dataframe(
        styled_df,
        use_container_width=True,
        hide_index=True,
        height=400,
        column_config=column_config
    )
    
    # Экспорт отфильтрованных данных
    if st.button("📥 Экспортировать отфильтрованные данные"):
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            filtered_df.to_excel(writer, sheet_name='Отфильтрованные отчеты', index=False)
            confirmation_reports.to_excel(writer, sheet_name='Требуют подтверждения', index=False)
            update_reports.to_excel(writer, sheet_name='Требуют актуализации', index=False)
        
        st.download_button(
            label="📁 Скачать аналитику",
            data=output.getvalue(),
            file_name=f"analytics_reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

def show_ai_assistant():
    st.markdown('<div class="page-header">❓ Частые вопросы</div>', unsafe_allow_html=True)
    st.markdown('<div class="coming-soon">🚧 Здесь будет интеграция с ИИ-ассистентом для ответов на вопросы</div>', unsafe_allow_html=True)

def show_feedback():
    st.markdown('<div class="page-header">💬 Оставить обратную связь</div>', unsafe_allow_html=True)
    st.markdown('<div class="coming-soon">🚧 Здесь будет форма для отправки обратной связи и предложений</div>', unsafe_allow_html=True)

def show_admin_control():
    st.markdown('<div class="page-header">🔍 Контроль публикации отчетов</div>', unsafe_allow_html=True)
    st.markdown('<span class="admin-badge">АДМИН</span>', unsafe_allow_html=True)
    
    st.markdown("## 📊 Анализатор запросов и стадий рассмотрения")
    st.info("Загрузите файл с запросами для анализа времени рассмотрения и текущих стадий")
    
    display_request_analysis()

def calculate_business_days(start_date, end_date):
    """Вычисляет количество рабочих дней между двумя датами по российскому производственному календарю"""
    if pd.isna(start_date) or pd.isna(end_date):
        return 0
    
    try:
        # Простой расчет рабочих дней (без workalendar для упрощения)
        # Исключаем выходные (суббота=5, воскресенье=6)
        current_date = start_date
        business_days = 0
        
        while current_date <= end_date:
            if current_date.weekday() < 5:  # Понедельник=0, Пятница=4
                business_days += 1
            current_date += timedelta(days=1)
        
        return business_days
    except Exception as e:
        return 0

def process_request_data(df):
    """Обработка данных согласно требованиям"""
    
    # Конвертируем даты
    date_columns = ['created_at', 'ts_from', 'ts_to']
    for col in date_columns:
        if col in df.columns:
            try:
                df[col] = pd.to_datetime(df[col], format='%d.%m.%Y', errors='coerce')
            except:
                try:
                    df[col] = pd.to_datetime(df[col], format='%Y-%m-%d', errors='coerce')
                except:
                    df[col] = pd.to_datetime(df[col], errors='coerce')
    
    # Сортируем по created_at от новых к старым
    df_sorted = df.sort_values('created_at', ascending=False)
    
    # Получаем уникальные запросы по business_id
    unique_requests = df_sorted.drop_duplicates(subset='business_id', keep='first')
    
    # Для каждого business_id находим последнюю строку для расчета дней в работе
    latest_records = df.groupby('business_id').apply(
        lambda x: x.loc[x['ts_from'].idxmax()] if x['ts_from'].notna().any() else x.iloc[-1]
    ).reset_index(drop=True)
    
    # Создаем итоговую таблицу
    result_data = []
    
    for _, unique_row in unique_requests.iterrows():
        business_id = unique_row['business_id']
        
        # Находим соответствующую последнюю запись для расчета дней
        latest_row = latest_records[latest_records['business_id'] == business_id].iloc[0]
        
        # Рассчитываем дни в работе (рабочие дни)
        if pd.notna(latest_row['ts_from']):
            days_in_work = calculate_business_days(latest_row['ts_from'], datetime.now())
        else:
            days_in_work = 0
        
        result_data.append({
            'business_id': int(business_id),
            'created_at': unique_row['created_at'].strftime('%d.%m.%Y') if pd.notna(unique_row['created_at']) else '',
            'рабочих_дней_в_работе': days_in_work,
            'form_type_report': unique_row.get('form_type_report', ''),
            'report_code': unique_row.get('report_code', ''),
            'report_name': unique_row.get('report_name', ''),
            'current_stage': unique_row.get('current_stage', ''),
            'ts_from': latest_row['ts_from'].strftime('%d.%m.%Y') if pd.notna(latest_row['ts_from']) else '',
            'analyst': unique_row.get('Analyst', ''),
            'request_owner': unique_row.get('request_owner', ''),
            'request_owner_ssp': unique_row.get('request_owner_ssp', '')
        })
    
    return pd.DataFrame(result_data)

def create_excel_download_requests(df):
    """Создание Excel файла для скачивания запросов"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Анализ запросов', index=False)
        
        # Получаем workbook и worksheet для форматирования
        workbook = writer.book
        worksheet = writer.sheets['Анализ запросов']
        
        # Автоподбор ширины столбцов
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output.getvalue()

def display_request_analysis():
    """Отображение анализатора запросов"""
    
    # Инициализация session state для анализатора (если не инициализированы)
    if 'request_processed_data' not in st.session_state:
        st.session_state.request_processed_data = None
    if 'request_original_data' not in st.session_state:
        st.session_state.request_original_data = None
    
    # Показываем статус загруженных данных
    if st.session_state.request_processed_data is not None:
        st.success(f"✅ Данные анализатора загружены: {len(st.session_state.request_processed_data)} запросов")
        
        # Показываем информацию о файлах
        col1, col2 = st.columns(2)
        with col1:
            if REQUESTS_DATA_FILE.exists():
                file_time = datetime.fromtimestamp(REQUESTS_DATA_FILE.stat().st_mtime)
                st.info(f"📁 Исходные данные: {file_time.strftime('%d.%m.%Y %H:%M')}")
        
        with col2:
            if REQUESTS_PROCESSED_FILE.exists():
                file_time = datetime.fromtimestamp(REQUESTS_PROCESSED_FILE.stat().st_mtime)
                st.info(f"🔄 Обработанные данные: {file_time.strftime('%d.%m.%Y %H:%M')}")
    else:
        st.info("📊 Данные анализатора не загружены")
    
    # Форма загрузки файла
    st.markdown("---")
    st.subheader("📁 Загрузка нового файла для анализа")
    
    uploaded_file = st.file_uploader(
        "Выберите файл с данными о запросах",
        type=['csv', 'xlsx'],
        help="Поддерживаются файлы в форматах CSV, XLSX. При загрузке нового файла предыдущие данные будут заменены.",
        key="request_analysis_uploader"
    )
    
    # Автоматическая загрузка и анализ файла
    if uploaded_file is not None:
        try:
            # Определяем тип файла и читаем соответствующим образом
            file_extension = uploaded_file.name.split('.')[-1].lower()
            
            if file_extension == 'csv':
                df = pd.read_csv(uploaded_file, encoding='utf-8')
            elif file_extension == 'xlsx':
                df = pd.read_excel(uploaded_file)
            else:
                st.error("❌ Неподдерживаемый формат файла!")
                return
            
            # Удаляем полностью пустые строки
            df = df.dropna(how='all')
            
            # Удаляем строки где business_id пустой
            if 'business_id' in df.columns:
                df = df.dropna(subset=['business_id'])
            else:
                st.error("❌ В файле отсутствует столбец 'business_id'")
                return
            
            st.session_state.request_original_data = df
            st.success(f"✅ Файл успешно загружен! Найдено {len(df)} записей.")
            
            # Автоматически обрабатываем данные
            try:
                processed_data = process_request_data(df)
                st.session_state.request_processed_data = processed_data
                
                # Сохраняем в постоянные файлы
                if save_requests_data(df, processed_data):
                    st.success("✅ Данные успешно обработаны и сохранены!")
                else:
                    st.error("❌ Ошибка при сохранении данных")
                
            except Exception as e:
                st.error(f"❌ Ошибка при обработке данных: {str(e)}")
            
        except Exception as e:
            st.error(f"❌ Ошибка при загрузке файла: {str(e)}")
    
    # Кнопка для принудительной загрузки из файлов
    if st.session_state.request_processed_data is None:
        if st.button("🔄 Попробовать загрузить данные из сохраненных файлов", key="req_load_from_files"):
            original_df, processed_df = load_requests_data()
            if processed_df is not None:
                st.session_state.request_original_data = original_df
                st.session_state.request_processed_data = processed_df
                st.success("✅ Данные успешно загружены из файлов!")
                st.rerun()
            else:
                st.error("❌ Сохраненные файлы не найдены. Загрузите новый файл.")
    
    # Отображение результатов
    if st.session_state.request_processed_data is not None:
        display_request_results(st.session_state.request_processed_data)
    
    # Управление данными (только если есть данные)
    if st.session_state.request_processed_data is not None:
        st.markdown("---")
        st.markdown("### 🗂️ Управление данными анализатора")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Информация о файлах:**")
            if REQUESTS_DATA_FILE.exists():
                file_size = REQUESTS_DATA_FILE.stat().st_size / 1024  # в KB
                st.text(f"📊 Исходные данные: {file_size:.1f} KB")
            
            if REQUESTS_PROCESSED_FILE.exists():
                file_size = REQUESTS_PROCESSED_FILE.stat().st_size / 1024  # в KB  
                st.text(f"🔄 Обработанные данные: {file_size:.1f} KB")
        
        with col2:
            if st.button("🔄 Очистить данные анализатора", key="req_clear_data", help="Удалит все данные анализатора запросов"):
                if st.session_state.get('confirm_clear_requests', False):
                    # Очищаем session state
                    st.session_state.request_original_data = None
                    st.session_state.request_processed_data = None
                    
                    # Удаляем файлы
                    try:
                        if REQUESTS_DATA_FILE.exists():
                            REQUESTS_DATA_FILE.unlink()
                        if REQUESTS_PROCESSED_FILE.exists():
                            REQUESTS_PROCESSED_FILE.unlink()
                        
                        st.success("✅ Данные анализатора очищены!")
                        st.session_state.confirm_clear_requests = False
                    except Exception as e:
                        st.error(f"❌ Ошибка при очистке: {str(e)}")
                    
                    st.rerun()
                else:
                    st.session_state.confirm_clear_requests = True
                    st.warning("⚠️ Нажмите еще раз для подтверждения очистки")
                    st.rerun()

def display_request_results(df):
    """Отображение результатов анализа запросов с фильтрами и поиском"""
    
    st.markdown("---")
    st.subheader("🔍 Результаты анализа")
    
    # Строка поиска
    st.subheader("🔎 Поиск")
    search_query = st.text_input(
        "Поиск по номеру отчета (report_code) или business_id:",
        placeholder="Введите номер отчета или business_id...",
        key="request_search"
    )
    
    # Применяем поиск
    filtered_df = df.copy()
    if search_query:
        search_mask = (
            filtered_df['report_code'].astype(str).str.contains(search_query, case=False, na=False) |
            filtered_df['business_id'].astype(str).str.contains(search_query, case=False, na=False)
        )
        filtered_df = filtered_df[search_mask]
    
    # Фильтры
    st.subheader("🔧 Фильтры")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        form_types = ['Все'] + sorted(df['form_type_report'].dropna().unique().tolist())
        selected_form_type = st.selectbox("Тип отчета:", form_types, key="req_form_type")
        
        analysts = ['Все'] + sorted(df['analyst'].dropna().unique().tolist())
        selected_analyst = st.selectbox("Аналитик:", analysts, key="req_analyst")
    
    with col2:
        stages = ['Все'] + sorted(df['current_stage'].dropna().unique().tolist())
        selected_stage = st.selectbox("Текущая стадия:", stages, key="req_stage")
        
        owners = ['Все'] + sorted(df['request_owner'].dropna().unique().tolist())
        selected_owner = st.selectbox("Владелец запроса:", owners, key="req_owner")
    
    with col3:
        owner_ssps = ['Все'] + sorted(df['request_owner_ssp'].dropna().unique().tolist())
        selected_owner_ssp = st.selectbox("Владелец ССП:", owner_ssps, key="req_owner_ssp")
        
        min_days = st.number_input("Мин. рабочих дней:", min_value=0, value=0, key="req_min_days")
    
    with col4:
        max_days = st.number_input("Макс. рабочих дней:", min_value=0, value=1000, key="req_max_days")
        
        # Кнопка сброса фильтров
        if st.button("🔄 Сбросить фильтры", key="req_reset_filters"):
            st.rerun()
    
    # Применяем фильтры
    if selected_form_type != 'Все':
        filtered_df = filtered_df[filtered_df['form_type_report'] == selected_form_type]
    
    if selected_stage != 'Все':
        filtered_df = filtered_df[filtered_df['current_stage'] == selected_stage]
    
    if selected_analyst != 'Все':
        filtered_df = filtered_df[filtered_df['analyst'] == selected_analyst]
    
    if selected_owner != 'Все':
        filtered_df = filtered_df[filtered_df['request_owner'] == selected_owner]
    
    if selected_owner_ssp != 'Все':
        filtered_df = filtered_df[filtered_df['request_owner_ssp'] == selected_owner_ssp]
    
    filtered_df = filtered_df[
        (filtered_df['рабочих_дней_в_работе'] >= min_days) & 
        (filtered_df['рабочих_дней_в_работе'] <= max_days)
    ]
    
    # Статистика
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("📊 Всего записей", len(df))
    with col2:
        st.metric("🔍 После фильтрации", len(filtered_df))
    with col3:
        if len(filtered_df) > 0:
            avg_days = filtered_df['рабочих_дней_в_работе'].mean()
            st.metric("📅 Среднее рабочих дней", f"{avg_days:.1f}")
        else:
            st.metric("📅 Среднее рабочих дней", "0")
    with col4:
        if len(filtered_df) > 0:
            max_days_value = filtered_df['рабочих_дней_в_работе'].max()
            st.metric("⏰ Максимум рабочих дней", max_days_value)
        else:
            st.metric("⏰ Максимум рабочих дней", "0")
    
    # Отображение таблицы
    st.subheader("📋 Таблица данных")
    
    if len(filtered_df) > 0:
        # Настройка отображения столбцов
        column_config = {
            'business_id': st.column_config.NumberColumn('business_id', format='%d', width="medium"),
            'created_at': st.column_config.TextColumn('Дата создания', width="medium"),
            'рабочих_дней_в_работе': st.column_config.NumberColumn('Рабочих дней в работе', format='%d', width="medium"),
            'form_type_report': st.column_config.TextColumn('Тип отчета', width="medium"),
            'report_code': st.column_config.TextColumn('Код отчета', width="medium"),
            'report_name': st.column_config.TextColumn('Название отчета', width="medium"),
            'current_stage': st.column_config.TextColumn('Текущая стадия', width="medium"),
            'ts_from': st.column_config.TextColumn('Дата начала', width="medium"),
            'analyst': st.column_config.TextColumn('Аналитик', width="medium"),
            'request_owner': st.column_config.TextColumn('Владелец запроса', width="medium"),
            'request_owner_ssp': st.column_config.TextColumn('Владелец ССП', width="medium")
        }
        
        st.dataframe(
            filtered_df,
            use_container_width=True,
            column_config=column_config,
            hide_index=True,
            height=400
        )
        
        # Кнопка экспорта в Excel
        excel_data = create_excel_download_requests(filtered_df)
        st.download_button(
            label="📥 Скачать результаты анализа",
            data=excel_data,
            file_name=f"requests_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="req_download"
        )
    else:
        st.warning("⚠️ Нет данных для отображения. Попробуйте изменить фильтры.")
    

def show_admin_dashboard():
    st.markdown('<div class="page-header">📊 Дашборд по отчетам</div>', unsafe_allow_html=True)
    st.markdown('<span class="admin-badge">АДМИН</span>', unsafe_allow_html=True)
    
    # Показываем статус загруженных данных
    if st.session_state.reports_data is not None:
        st.success(f"✅ Данные загружены: {len(st.session_state.reports_data)} отчетов")
        
        # Показываем информацию о файлах
        col1, col2 = st.columns(2)
        with col1:
            if REPORTS_DATA_FILE.exists():
                file_time = datetime.fromtimestamp(REPORTS_DATA_FILE.stat().st_mtime)
                st.info(f"📁 Файл данных: {file_time.strftime('%d.%m.%Y %H:%M')}")
        
        with col2:
            if COMMENTS_DATA_FILE.exists():
                file_time = datetime.fromtimestamp(COMMENTS_DATA_FILE.stat().st_mtime)
                st.info(f"💬 Комментарии: {file_time.strftime('%d.%m.%Y %H:%M')}")
    else:
        st.info("📊 Данные не загружены")
    
    # Загрузка нового файла
    st.markdown("---")
    st.markdown("## 📁 Загрузка новых данных отчетов")
    
    uploaded_file = st.file_uploader(
        "Выберите файл с данными отчетов",
        type=['xlsx', 'xls', 'csv'],
        key="reports_file_uploader",
        help="При загрузке нового файла комментарии сохранятся для совпадающих отчетов"
    )
    
    if uploaded_file is not None:
        try:
            # Загружаем новые данные
            if uploaded_file.name.endswith('.csv'):
                new_df = pd.read_csv(uploaded_file)
            else:
                new_df = pd.read_excel(uploaded_file)
            
            # Сохраняем старые комментарии
            old_comments = st.session_state.reports_comments.copy()
            
            # Обновляем данные
            st.session_state.reports_data = new_df
            
            # Пытаемся сопоставить старые комментарии с новыми данными
            new_comments = {}
            if not new_df.empty and old_comments:
                for old_idx, comment in old_comments.items():
                    # Ищем совпадения по номеру формы или названию отчета
                    if old_idx < len(st.session_state.reports_data):
                        old_report_info = st.session_state.reports_data.iloc[old_idx] if old_idx < len(st.session_state.reports_data) else None
                        if old_report_info is not None:
                            # Ищем тот же отчет в новых данных
                            for new_idx, row in new_df.iterrows():
                                if (str(row.get('Номер формы', '')) == str(old_report_info.get('Номер формы', '')) and
                                    str(row.get('Наименование отчета', '')) == str(old_report_info.get('Наименование отчета', ''))):
                                    new_comments[new_idx] = comment
                                    break
            
            st.session_state.reports_comments = new_comments
            
            # Сохраняем в постоянные файлы
            if save_reports_data(new_df, st.session_state.reports_comments):
                st.success("✅ Файл успешно загружен и сохранен!")
                
                preserved_comments = len(new_comments)
                if preserved_comments > 0:
                    st.info(f"💬 Сохранено {preserved_comments} комментариев из предыдущей версии")
            else:
                st.error("❌ Ошибка при сохранении данных")
            
            # Показываем базовую информацию
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Всего отчетов", len(new_df))
            with col2:
                st.metric("Столбцов данных", len(new_df.columns))
            with col3:
                owners_count = new_df['Владелец отчета ССП'].nunique() if 'Владелец отчета ССП' in new_df.columns else 0
                st.metric("Уникальных владельцев", owners_count)
        
        except Exception as e:
            st.error(f"❌ Ошибка при загрузке файла: {str(e)}")
    
    # Отображение данных если они есть
    if st.session_state.reports_data is not None:
        df = st.session_state.reports_data.copy()
        
        st.markdown("---")
        st.markdown("## 📋 Управление данными отчетов")
        
        # Фильтры
        col1, col2 = st.columns(2)
        with col1:
            if 'Владелец отчета ССП' in df.columns:
                owners = ['Все'] + sorted(df['Владелец отчета ССП'].dropna().unique().tolist())
                selected_owner = st.selectbox("Фильтр по владельцу ССП", owners)
            else:
                selected_owner = "Все"
        
        with col2:
            if 'Этап отчета' in df.columns:
                stages = ['Все'] + sorted(df['Этап отчета'].dropna().unique().tolist())
                selected_stage = st.selectbox("Фильтр по этапу отчета", stages)
            else:
                selected_stage = "Все"
        
        # Применяем фильтры
        filtered_df = df.copy()
        if selected_owner != "Все":
            filtered_df = filtered_df[filtered_df['Владелец отчета ССП'] == selected_owner]
        if selected_stage != "Все":
            filtered_df = filtered_df[filtered_df['Этап отчета'] == selected_stage]
        
        # Показываем количество записей
        st.markdown(f"**Отображается записей: {len(filtered_df)} из {len(df)}**")
        
        # Отображаем таблицу с подсветкой пустых ячеек
        def highlight_empty_cells(df_display):
            def style_func(val):
                if pd.isna(val) or str(val).strip() == '':
                    return 'background-color: #ffcccc'
                return ''
            return df_display.style.applymap(style_func)
        
        # Форматируем даты в админской таблице
        admin_display_df = filtered_df.copy()
        
        # Форматируем даты в дд.мм.гггг
        date_columns = ['Дата создания последнего черновика', 'Дата последней публикации отчета']
        for col in date_columns:
            if col in admin_display_df.columns:
                admin_display_df[col] = pd.to_datetime(admin_display_df[col], errors='coerce').dt.strftime('%d.%m.%Y')
                admin_display_df[col] = admin_display_df[col].replace('NaT', '')
        
        # Создаем конфигурацию для всех столбцов с одинаковой шириной
        admin_column_config = {}
        for col in admin_display_df.columns:
            admin_column_config[col] = st.column_config.TextColumn(col, width="small")
        
        styled_df = highlight_empty_cells(admin_display_df)
        st.dataframe(
            styled_df,
            use_container_width=True,
            hide_index=True,
            height=400,
            column_config=admin_column_config
        )
        
        # Управление комментариями
        st.markdown("---")
        st.markdown("### 💬 Управление комментариями")
        
        if not filtered_df.empty:
            # Выбор отчета для комментирования
            report_options = []
            report_indices = []
            
            for idx, row in filtered_df.iterrows():
                option_text = f"{row.get('Номер формы', 'N/A')} - {row.get('Наименование отчета', 'N/A')}"
                report_options.append(option_text)
                report_indices.append(idx)
            
            selected_report_idx = st.selectbox(
                "Выберите отчет для комментирования", 
                range(len(report_options)),
                format_func=lambda x: report_options[x]
            )
            
            if selected_report_idx is not None:
                actual_idx = report_indices[selected_report_idx]
                current_comment = st.session_state.reports_comments.get(actual_idx, '')
                
                new_comment = st.text_area(
                    "Комментарий", 
                    value=current_comment, 
                    height=100,
                    key="comment_editor"
                )
                
                col1, col2 = st.columns([1, 1])
                
                with col1:
                    if st.button("💾 Сохранить комментарий", type="primary"):
                        st.session_state.reports_comments[actual_idx] = new_comment
                        
                        # Сохраняем в постоянный файл
                        if save_reports_data(st.session_state.reports_data, st.session_state.reports_comments):
                            st.success("✅ Комментарий сохранен!")
                        else:
                            st.error("❌ Ошибка при сохранении комментария")
                        st.rerun()
                
                with col2:
                    if st.button("🗑️ Удалить комментарий"):
                        if actual_idx in st.session_state.reports_comments:
                            del st.session_state.reports_comments[actual_idx]
                            
                            # Сохраняем в постоянный файл
                            if save_reports_data(st.session_state.reports_data, st.session_state.reports_comments):
                                st.success("✅ Комментарий удален!")
                            else:
                                st.error("❌ Ошибка при удалении комментария")
                            st.rerun()
                
                # Показываем существующие комментарии
                comments_count = len([c for c in st.session_state.reports_comments.values() if c.strip()])
                if comments_count > 0:
                    st.info(f"📝 Всего комментариев в системе: {comments_count}")
        
        # Экспорт данных
        st.markdown("---")
        st.markdown("### 📥 Экспорт данных")
        
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("📊 Скачать отфильтрованные данные", type="primary"):
                # Добавляем комментарии в DataFrame
                export_df = filtered_df.copy()
                export_df['Комментарий'] = ''
                
                for idx in export_df.index:
                    if idx in st.session_state.reports_comments:
                        export_df.at[idx, 'Комментарий'] = st.session_state.reports_comments[idx]
                
                # Создаем Excel файл
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    export_df.to_excel(writer, sheet_name='Отфильтрованные отчеты', index=False)
                
                st.download_button(
                    label="📁 Скачать Excel файл",
                    data=output.getvalue(),
                    file_name=f"filtered_reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        with col2:
            if st.button("💾 Скачать все данные с комментариями"):
                # Добавляем все комментарии в DataFrame
                export_df = df.copy()
                export_df['Комментарий'] = ''
                
                for idx in export_df.index:
                    if idx in st.session_state.reports_comments:
                        export_df.at[idx, 'Комментарий'] = st.session_state.reports_comments[idx]
                
                # Создаем Excel файл
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    export_df.to_excel(writer, sheet_name='Все отчеты', index=False)
                
                st.download_button(
                    label="📁 Скачать полный Excel файл",
                    data=output.getvalue(),
                    file_name=f"all_reports_with_comments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-fficedocument.spreadsheetml.sheet"
                )

def show_admin_stats():
    st.markdown('<div class="page-header">📈 Статистика по публикации</div>', unsafe_allow_html=True)
    st.markdown('<span class="admin-badge">АДМИН</span>', unsafe_allow_html=True)
    st.markdown('<div class="coming-soon">🚧 Административная статистика и аналитика по публикациям</div>', unsafe_allow_html=True)

def show_admin_issues():
    st.markdown('<div class="page-header">⚠️ Проблемные вопросы</div>', unsafe_allow_html=True)
    st.markdown('<span class="admin-badge">АДМИН</span>', unsafe_allow_html=True)
    st.markdown('<div class="coming-soon">🚧 Мониторинг и управление проблемными вопросами</div>', unsafe_allow_html=True)

# Отображение выбранной страницы
# Определяем, какую страницу показывать - админ или основную
show_admin_page = False
if st.session_state.admin_mode and 'selected_admin_page' in st.session_state:
    # Проверяем, была ли нажата админ кнопка последней
    admin_pages = {
        "🔍 Контроль публикации отчетов": "admin_control",
        "📊 Дашборд по отчетам": "admin_dashboard", 
        "📈 Статистика по публикации": "admin_stats",
        "⚠️ Проблемные вопросы": "admin_issues"
    }
    
    # Если выбранная админ страница существует, показываем её
    if st.session_state.selected_admin_page in admin_pages.keys():
        show_admin_page = True

if show_admin_page:
    # Показываем админ страницы
    admin_page_map = {
        "🔍 Контроль публикации отчетов": show_admin_control,
        "📊 Дашборд по отчетам": show_admin_dashboard,
        "📈 Статистика по публикации": show_admin_stats, 
        "⚠️ Проблемные вопросы": show_admin_issues
    }
    admin_page_map[st.session_state.selected_admin_page]()
else:
    # Показываем основные страницы
    main_page_map = {
        "📈 Дашборд по отчетам": show_dashboard,
        "⚡ Действия с отчетами": show_actions,
        "🏷️ Сформировать атрибуты": show_attributes,
        "❓ Частые вопросы": show_ai_assistant,
        "📋 Документация": show_instructions,
        "💬 Оставить обратную связь": show_feedback
    }
    main_page_map[st.session_state.selected_page]()

# Подвал приложения
st.markdown("---")
col1, col2, col3 = st.columns(3)
with col1:
    st.caption("🏢 Компания XYZ")
with col2:
    st.caption("📧 support@company.com")
with col3:
    st.caption("📞 +7 (xxx) xxx-xx-xx")
